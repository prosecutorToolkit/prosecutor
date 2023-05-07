import openpyxl, os, shutil
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from message import success, error

def createExcelFile(path_xlsx, listOfData, headObj, ignoredFilesList):
    # try:

    book = openpyxl.Workbook()
    sheet_All = book.create_sheet(title="Scanned files")
    sheet_Match = book.create_sheet(title="Match files")
    sheet_Ignored = book.create_sheet(title="Ignored files")
    book.remove(book['Sheet'])

    # CREAR UN REPORTE DE ERRORES, DEBE ESTAR PARA TODOS LOS REPORTES WORD CSV EXCEL
    # sheet_Match.append((("IMAGE SEARCH ENGINE", "", "", row_errors)))

    try:
        imgPath = os.path.join(os.getcwd(), 'logo.jpg')
        imag = Image(imgPath)
        imag.width = 150
        imag.height = 150

        imgPath2 = os.path.join(os.getcwd(), 'logo2.jpg')
        shutil.copy(imgPath, imgPath2)
        imag2 = Image(imgPath2)
        imag2.width = 150
        imag2.height = 150

        sheet_All.add_image(imag, 'B2')
        sheet_Match.add_image(imag2, 'B2')
    except: pass
    
    # IGNORED FILES
    sheet_Ignored.row_dimensions[1].height = 10
    sheet_Ignored.column_dimensions['A'].width = 3
    sheet_Ignored.column_dimensions['B'].width = 6
    sheet_Ignored.column_dimensions['C'].width = 15
    sheet_Ignored.append((("", "FILE", "EXT", "PATH")))

    try:
        for idF, file, ext, path in ignoredFilesList:
            data = ((idF, file, ext, path))
            sheet_Ignored.append(data)
    except:
        idF = 0
        for file, ext, path in ignoredFilesList:
            idF += 1
            data = ((idF, file, ext, path))
            sheet_Ignored.append(data)

    # MATCH FILES
    sheet_Match.row_dimensions[1].height = 10
    sheet_Match.column_dimensions['A'].width = 3
    sheet_Match.column_dimensions['B'].width = 6
    sheet_Match.column_dimensions['C'].width = 15
    sheet_Match.column_dimensions['D'].width = 40
    sheet_Match.column_dimensions['E'].width = 25
    sheet_Match.column_dimensions['G'].width = 40
    sheet_Match.column_dimensions['H'].width = 40
    sheet_Match.append((("", "")))
    sheet_Match.append((("", "", "", "MATCH FILES")))
    sheet_Match.append((("", "", "", headObj.user)))
    sheet_Match.append((("", "", "", headObj.case)))
    sheet_Match.append((("", "", "", headObj.time)))
    sheet_Match.append((("", "", "", headObj.target)))
    sheet_Match.append((("", "", "", headObj.mercle)))
    sheet_Match.append((("", "", "", headObj.formats)))
    sheet_Match.append((("", "", "", headObj.terms)))
    sheet_Match.append((("", "", "", headObj.misspellings)))
    sheet_Match.append((("", "PROSECUTOR REPORT")))
    sheet_Match.append((("   ___________________________________________________________________________________________________________________________________________________________________________________________________________________", "")))
    sheet_Match.append((("", "ID", "NAME", "DIRECTION", "HASH SHA-256", "MATCH", "CUT TEXT", "METADATA")))

    sheet_Match['D2'].font = Font(bold=True)

    # SCANNED FILES
    sheet_All.row_dimensions[1].height = 10
    sheet_All.column_dimensions['A'].width = 3
    sheet_All.column_dimensions['B'].width = 6
    sheet_All.column_dimensions['C'].width = 15
    sheet_All.column_dimensions['D'].width = 40
    sheet_All.column_dimensions['E'].width = 25
    sheet_All.column_dimensions['G'].width = 40
    sheet_All.column_dimensions['H'].width = 40

    sheet_All.append((("", "")))
    sheet_All.append((("", "", "", "ALL IMAGES")))
    sheet_All.append((("", "", "", headObj.user)))
    sheet_All.append((("", "", "", headObj.case)))
    sheet_All.append((("", "", "", headObj.time)))
    sheet_All.append((("", "", "", headObj.target)))
    sheet_All.append((("", "", "", headObj.mercle)))
    sheet_All.append((("", "", "", headObj.formats)))
    sheet_All.append((("", "")))
    sheet_All.append((("", "PROSECUTOR REPORT")))
    sheet_All.append((("___________________________________________________________________________________________________________________________________________________________________________________________________________________", "")))

    sheet_All.append((("", "ID", "NAME", "DIRECTION", "HASH SHA-256", "MATCH", "CUT TEXT", "METADATA")))
    bold_cells = sheet_All["B11":"G11"]

    font = Font(bold=True)
    for row in bold_cells:
        for cell in row: cell.font = font

    sheet_All['D2'].font = Font(bold=True)
    sheet_All['B9'].font = Font(bold=True) 

    for idF, name, file, hashF, match, cutText, text, metadata in listOfData:
        data = (("", idF, name, file, hashF, match, cutText, metadata))
        sheet_All.append(data)
        if len(match) > 2: sheet_Match.append(data)
    
    book.save(path_xlsx)

    success("Xlsx was created")

    return True

    # except:
    #     error("reated")
    #     return False